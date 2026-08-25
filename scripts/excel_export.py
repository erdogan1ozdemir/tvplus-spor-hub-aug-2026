#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Keyword veri setinin tam Excel çıktısı.

Mimari: Özet sayfası + konu başına detay sayfası. Özet toplamları Python
sabitleriyle değil SUMIF/COUNTIF formülleriyle kurulur; böylece detay
sayfasında filtre uygulandığında da tutarlı kalır.

Renk sistemi Inbound Design System'in Excel katmanına uyar: başlık satırı
Shadowed Charcoal, gövde Calibri + ink teal, değişim sütunlarında yalnızca
yazı rengi (dolgu yok), durum sütununda dörtlü rozet paleti.
"""
import json, sys
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ——— Renk ve tipografi
BASLIK_BG = "FF434343"; BASLIK_YAZI = "FFFFFFFF"
INK       = "FF10332F"; CIZGI = "FFE0E0E0"
YESIL     = "FF2E7D32"; KIRMIZI = "FFD32F2F"; CORAL = "FFFF7B52"
DURUM = {  # (dolgu, yazı)
    "Mevcut":     ("FFC8E6C9", "FF2E7D32"),
    "Düzeltildi": ("FFF5A623", "FF10332F"),
    "Eksik":      ("FFFFCDD2", "FFD32F2F"),
    "Planlandı":  ("FFFFE3D8", "FFE85F36"),
}
GOVDE   = Font(name="Calibri", size=10, color=INK)
BAS_F   = Font(name="Calibri", size=10, bold=True, color=BASLIK_YAZI)
BAS_FILL= PatternFill("solid", fgColor=BASLIK_BG)
UST_KEN = Border(top=Side(style="thin", color=CIZGI))
ORTA    = Alignment(horizontal="center", vertical="center", wrap_text=False)
SOL     = Alignment(horizontal="left",   vertical="center")
BAS_HIZ = Alignment(horizontal="center", vertical="center", wrap_text=True)

def yaz_baslik(ws, basliklar, satir=1):
    for j, b in enumerate(basliklar, 1):
        c = ws.cell(satir, j, b); c.font = BAS_F; c.fill = BAS_FILL; c.alignment = BAS_HIZ
    ws.row_dimensions[satir].height = 30

def bicimle(ws, basliklar, ilk_veri, son_veri, metin_kolon=(), delta_kolon=(),
            durum_kolon=(), sayi_bicim=None):
    """Gövde biçimi: dikeyde ortalı, metin solda, sayı ortada, delta yalnız renkli."""
    for i in range(ilk_veri, son_veri + 1):
        for j in range(1, len(basliklar) + 1):
            c = ws.cell(i, j)
            c.font = GOVDE
            c.alignment = SOL if j in metin_kolon else ORTA
            c.border = UST_KEN
            if sayi_bicim and j in sayi_bicim: c.number_format = sayi_bicim[j]
            if j in delta_kolon and isinstance(c.value, (int, float)):
                # Değişim sütunlarında dolgu kullanılmaz, yalnızca yazı rengi
                c.font = Font(name="Calibri", size=10, bold=True,
                              color=YESIL if c.value > 0 else KIRMIZI if c.value < 0 else INK)
            if j in durum_kolon and c.value in DURUM:
                bg, fg = DURUM[c.value]
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(name="Calibri", size=10, bold=True, color=fg)

def genislik(ws, genislikler):
    for j, w in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

# ————————————————————————————————————————————————— veri
D = json.load(open("/tmp/dash.json", encoding="utf-8"))
M, KW, AYLAR = D["meta"], D["keywords"], D["meta"]["aylar"]
FASET = D["facetAdlari"]                       # csv adı → kısa ad
TERS  = {v: k for k, v in FASET.items()}
wb = Workbook(); wb.remove(wb.active)

FASET_SIRA = ["spor","org","mus","sev","cins","km","tb","cog","yer","turk","hak",
              "per","tak","st","it","ent","marka","dil","uzn","ktm","kulup",
              "anaAd","kurum","mden"]
FASET_ETIKET = {
 "spor":"Spor Dalı","org":"Organizasyon","mus":"Müsabaka Tipi","sev":"Lig Seviyesi",
 "cins":"Cinsiyet","km":"Kulüp / Milli","tb":"Takım / Bireysel","cog":"Coğrafya",
 "yer":"Yerlilik","turk":"Türk Bağlantısı","hak":"Yayın Hakkı","per":"Periyodiklik",
 "tak":"Takvim Tipi","st":"Sayfa Tipi","it":"Intent Katmanı","ent":"Varlık Tipi",
 "marka":"Marka Tipi","dil":"Dil","uzn":"Sorgu Uzunluğu","ktm":"Katman",
 "kulup":"Kulüp","anaAd":"Oyuncu Ana Adı","kurum":"Kurum Sorgusu","mden":"Mantık Denetimi"}

def ay_toplam(k, alan): return sum(v or 0 for v in (k.get(alan) or []))

# ———————————————————————————————————— 1) Keyword Listesi
ws = wb.create_sheet("Keyword Listesi")
BAS = (["Keyword"] + [FASET_ETIKET[f] for f in FASET_SIRA] +
       ["Son 12 Ay","Önceki 12 Ay","Rolling YoY",
        M["yillar"][0],M["yillar"][1],M["yillar"][2],
        "Takvim YoY","YTD YoY",
        f'{M["yillar"][0]} Aylık Ort.',f'{M["yillar"][1]} Aylık Ort.',f'{M["yillar"][2]} Aylık Ort.',
        "Hacim Bandı","Mevsim Tipi","CV","Peak / Dip","Trend","Peak Ayı","Dip Ayı","Kaynak"])
yaz_baslik(ws, BAS)
for k in KW:
    ws.append([k.get("kw")] + [k.get(f) or "" for f in FASET_SIRA] + [
        k.get("r12") or 0, k.get("p12") or 0, k.get("ryoy"),
        ay_toplam(k,"m24"), ay_toplam(k,"m25"), ay_toplam(k,"m26"),
        k.get("yoy"), k.get("ytd"),
        k.get("a24"), k.get("a25"), k.get("a26"),
        k.get("bucket") or "", k.get("sinif") or "", k.get("cv"), k.get("pd"),
        k.get("trend") or "", k.get("peakYm") or "", k.get("dipYm") or "", k.get("kaynak") or ""])
son = ws.max_row
n = len(BAS)
i_r12 = len(FASET_SIRA) + 2
delta = {i_r12+2, i_r12+6, i_r12+7}                       # Rolling YoY, Takvim YoY, YTD YoY
sayi  = {j: "#,##0" for j in (i_r12, i_r12+1, i_r12+3, i_r12+4, i_r12+5,
                              i_r12+8, i_r12+9, i_r12+10)}
sayi.update({j: "0.0%" for j in delta})
sayi[i_r12+12+1] = "0.00"; sayi[i_r12+12+2] = "0.0"
bicimle(ws, BAS, 2, son, metin_kolon={1,2,3}, delta_kolon=delta, sayi_bicim=sayi)
genislik(ws, [34] + [15]*len(FASET_SIRA) + [13]*(n-len(FASET_SIRA)-1))
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(n)}{son}"
print(f"Keyword Listesi : {son-1} satır × {n} sütun")

# Keyword Listesi sütun harfleri (özet formülleri buraya bağlanır)
KOL = {f: get_column_letter(2 + i) for i, f in enumerate(FASET_SIRA)}
K_R12, K_P12 = get_column_letter(i_r12), get_column_letter(i_r12 + 1)
KL = "'Keyword Listesi'"

def ozet_sayfa(ad, faset, baslik_ad):
    """Bir faset için özet: toplamlar SUMIF, adetler COUNTIF ile kurulur."""
    w = wb.create_sheet(ad)
    bas = [baslik_ad, "Keyword", "Son 12 Ay", "Önceki 12 Ay", "YoY", "Pay"]
    yaz_baslik(w, bas)
    degerler = sorted({(k.get(faset) or "–") for k in KW})
    kol = KOL[faset]
    for d in degerler:
        w.append([d,
            f'=COUNTIF({KL}!${kol}:${kol},$A{w.max_row+1})',
            f'=SUMIF({KL}!${kol}:${kol},$A{w.max_row+1},{KL}!${K_R12}:${K_R12})',
            f'=SUMIF({KL}!${kol}:${kol},$A{w.max_row+1},{KL}!${K_P12}:${K_P12})',
            f'=IF($D{w.max_row+1}>0,($C{w.max_row+1}-$D{w.max_row+1})/$D{w.max_row+1},"")',
            f'=IF(SUM($C$2:$C${len(degerler)+1})>0,$C{w.max_row+1}/SUM($C$2:$C${len(degerler)+1}),"")'])
    son = w.max_row
    t = son + 1
    w.append(["Toplam", f"=SUM(B2:B{son})", f"=SUM(C2:C{son})", f"=SUM(D2:D{son})",
              f'=IF($D{t}>0,($C{t}-$D{t})/$D{t},"")', f"=SUM(F2:F{son})"])
    bicimle(w, bas, 2, t, metin_kolon={1}, delta_kolon={5},
            sayi_bicim={2:"#,##0", 3:"#,##0", 4:"#,##0", 5:"0.0%", 6:"0.0%"})
    for j in range(1, 7):
        c = w.cell(t, j); c.font = Font(name="Calibri", size=10, bold=True, color=INK)
        c.border = Border(top=Side(style="medium", color=CIZGI))
    genislik(w, [30, 12, 16, 16, 11, 10]); w.freeze_panes = "A2"
    return son

for ad, faset, bas in [("Spor Dalı","spor","Spor Dalı"), ("Organizasyon","org","Organizasyon"),
                       ("Sayfa Tipi","st","Sayfa Tipi"), ("Intent","it","Intent Katmanı"),
                       ("Varlık Tipi","ent","Varlık Tipi"), ("Yayın Hakkı","hak","Yayın Hakkı"),
                       ("Katman","ktm","Katman")]:
    n_ = ozet_sayfa(ad, faset, bas)
    print(f"{ad:<16}: {n_-1} satır")

# ———————————————————————————————————— Aylık Seri
ws = wb.create_sheet("Aylık Seri")
bas = ["Keyword", "Spor Dalı", "Organizasyon", "Varlık Tipi"] + AYLAR
yaz_baslik(ws, bas)
for k in KW:
    seri = (k.get("m24") or []) + (k.get("m25") or []) + (k.get("m26") or [])
    seri = (seri + [None] * len(AYLAR))[:len(AYLAR)]
    ws.append([k.get("kw"), k.get("spor") or "", k.get("org") or "", k.get("ent") or ""] + seri)
son = ws.max_row
bicimle(ws, bas, 2, son, metin_kolon={1,2,3},
        sayi_bicim={j: "#,##0" for j in range(5, len(bas)+1)})
genislik(ws, [34, 15, 20, 13] + [10]*len(AYLAR))
ws.freeze_panes = "E2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(bas))}{son}"
print(f"Aylık Seri      : {son-1} satır × {len(AYLAR)} ay")

# ———————————————————————————————————— Denetim (dashboard dışında bırakılan satırlar)
import csv, glob
isaretli = []
for d in [x for x in sorted(glob.glob("data/raw/hacim_*.csv")) if not x.endswith("_elenen.csv")]:
    for r in csv.DictReader(open(d, encoding="utf-8-sig")):
        for kol, tip in (("mantik_denetim","Mantık"), ("varyant_denetim","Varyant")):
            v = (r.get(kol) or "").strip()
            if v and v != "Geçerli":
                isaretli.append([r.get("keyword"), tip, v, r.get("spor_dali") or "",
                                 r.get("organizasyon") or "", r.get("entity_tipi") or "",
                                 int(r.get("search_volume") or 0) if r.get("veri_var")=="evet" else 0,
                                 "Eksik"])
ws = wb.create_sheet("Denetim")
bas = ["Keyword","Denetim","Bulgu","Spor Dalı","Organizasyon","Varlık Tipi","Aylık Hacim","Durum"]
yaz_baslik(ws, bas)
if isaretli:
    for r in sorted(isaretli, key=lambda x: -x[6]): ws.append(r)
else:
    ws.append(["Bu taramada denetim bulgusu tespit edilmedi.", "", "", "", "", "", "", ""])
son = ws.max_row
bicimle(ws, bas, 2, son, metin_kolon={1,3,5}, durum_kolon={8}, sayi_bicim={7:"#,##0"})
genislik(ws, [30,12,34,15,24,14,14,13]); ws.freeze_panes = "A2"
# Öneri sheet'in yalnızca ilk veri satırına yazılır (tablo veri, öneri kural)
o = ws.cell(2, len(bas)+1, "B sütunundaki denetim, A sütunundaki keyword'ün ilgili "
    "faset kuralını karşılamadığını göstermektedir; bu satırlar dashboard "
    "toplamlarına dahil edilmemiştir, kaynak veri setinde korunmaktadır.")
o.font = Font(name="Calibri", size=10, italic=True, color=INK); o.alignment = SOL
ws.column_dimensions[get_column_letter(len(bas)+1)].width = 90
print(f"Denetim         : {len(isaretli)} satır")

# ———————————————————————————————————— Özet (en başa alınır)
ws = wb.create_sheet("Özet")
ws["A1"] = (f'Turkcell TV+ Spor Keyword Veri Seti  |  Kaynak: {M["kaynak"]}  |  '
            f'Dönem: {AYLAR[0]} - {AYLAR[-1]} ({len(AYLAR)} ay)  |  Oluşturma: {M["olusturma"]}')
ws["A1"].font = Font(name="Calibri", size=11, bold=True, color=INK)
ws.merge_cells("A1:F1"); ws.row_dimensions[1].height = 22

bas = ["Gösterge", "Değer", "Açıklama"]
yaz_baslik(ws, bas, satir=3)
satirlar = [
 ("Toplam keyword", f"=COUNTA({KL}!$A$2:$A${len(KW)+1})", "Denetimden geçen, dashboard'da yer alan satır sayısı"),
 ("Son 12 Ay arama", f"=SUM({KL}!${K_R12}$2:${K_R12}${len(KW)+1})", f'{D["monthsR12"][0]} - {D["monthsR12"][-1]}'),
 ("Önceki 12 Ay arama", f"=SUM({KL}!${K_P12}$2:${K_P12}${len(KW)+1})", "Bir önceki 12 aylık pencere"),
 ("Rolling YoY", "=IF($B5>0,($B4-$B5)/$B5,\"\")", "Son 12 Ay ile Önceki 12 Ay karşılaştırması"),
 ("Spor dalı sayısı", "=COUNTA('Spor Dalı'!$A$2:$A$400)-1", "Detay: Spor Dalı sayfası"),
 ("Organizasyon sayısı", "=COUNTA(Organizasyon!$A$2:$A$400)-1", "Detay: Organizasyon sayfası"),
 ("Denetimde işaretli satır", f"=COUNTA(Denetim!$A$2:$A${max(len(isaretli)+1,2)})", "Toplamlara dahil edilmemiştir"),
]
for s in satirlar: ws.append([]) if False else ws.append(list(s))
son = ws.max_row
bicimle(ws, bas, 4, son, metin_kolon={1,3}, delta_kolon={2} if False else set(),
        sayi_bicim={2:"#,##0"})
ws.cell(6, 2).number_format = "0.0%"
genislik(ws, [30, 20, 62])

# ———————————————————————————————————— Yöntem
ws = wb.create_sheet("Yöntem")
notlar = [
 ("Veri kaynağı", M["kaynak"]),
 ("Dönem", f'{AYLAR[0]} - {AYLAR[-1]} ({len(AYLAR)} aylık seri)'),
 ("Son 12 Ay", f'{D["monthsR12"][0]} - {D["monthsR12"][-1]}'),
 ("Takvim yılları", " / ".join(str(y) for y in M["yillar"])),
 ("Rolling YoY", "Son 12 aylık toplamın önceki 12 aylık toplamla karşılaştırması"),
 ("Takvim YoY", "İki tam takvim yılının karşılaştırması"),
 ("YTD YoY", "Kısmi son yılın, önceki yılın aynı ay sayısıyla karşılaştırması"),
 ("Mevsim tipi", "Tüm aylık seri üzerinden: CV 0.35 altı Evergreen, peak/dip 20 ve üzeri "
                 "veya CV 1.0 ve üzeri Spike, kalanlar Seasonal"),
 ("Trend", "Rolling YoY +%5 üzeri Yükselen, -%5 altı Düşen, arası Stabil"),
 ("Hacim bandı", "Son 12 Ay toplamının aylık ortalamasına göre gruplanır"),
 ("Denetim", "Faset kurallarını karşılamayan satırlar Denetim sayfasında listelenir ve "
             "toplamlara dahil edilmez; kaynak veri setinden silinmez"),
 ("Kapsam notu", "Hacimler Türkiye ve Türkçe için alınmıştır. Google Ads benzer yazımları "
                 "tek hacimde birleştirdiğinden tekil ve çoğul varyantlar teke indirilmiştir"),
]
bas = ["Başlık", "Açıklama"]; yaz_baslik(ws, bas)
for a, b in notlar: ws.append([a, b])
son = ws.max_row
bicimle(ws, bas, 2, son, metin_kolon={1,2})
for i in range(2, son+1):
    c = ws.cell(i, 1)
    c.font = Font(name="Calibri", size=10, bold=True, color=CORAL)
    c.border = Border(top=Side(style="thin", color=CIZGI),
                      left=Side(style="thin", color=CORAL))
    ws.cell(i, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
genislik(ws, [24, 96])

# Okuma sırası: önce Özet ve Yöntem, sonra özet kırılımları, en sonda ham detay
SIRA = ["Özet","Yöntem","Spor Dalı","Organizasyon","Sayfa Tipi","Intent","Varlık Tipi",
        "Yayın Hakkı","Katman","Denetim","Keyword Listesi","Aylık Seri"]
wb._sheets = [wb[a] for a in SIRA if a in wb.sheetnames] + \
             [w for w in wb._sheets if w.title not in SIRA]
wb.active = 0
cikti = "data/tvplus-spor-keyword-veri-seti.xlsx"
wb.save(cikti)
import os
print(f"\nÇıktı: {cikti}  ({os.path.getsize(cikti)/1048576:.1f} MB)")
print("Sayfalar:", " · ".join(wb.sheetnames))
